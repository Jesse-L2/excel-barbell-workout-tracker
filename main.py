"""
main.py
~~~~~~~
The main function of this program uses openpyxl to read the intended weight to be
lifted from various locations in the 'Workout.xlsx' spreadsheet and then apply and
output the weight_calc() function from weight_calculator.py to the various rows and
columns of the 'Workout.xlsx' spreadsheet

Note - 'Workout.xlsx' is a requirement for this program to run, however weight_calc()
will function if called separately onto any int or floating point number, provided that
number is less than the total sum of weights within the weight_calc function in
weight_calculator.py

"""
# TODO: change from openpyxl to xlwings https://docs.xlwings.org/en/stable/quickstart.html
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import weight_calculator


def iter_weight_calc(row_min, row_max, col, output_col, sheet):
    """Takes the value from each column in the range between row_min and row_max - 1 and iterate
    over each equivalent range of output_col with weight_calc() from weight_calculator.py and clean
    up the output"""
    for i in range(row_min, row_max):
        total = sheet.cell(row=i, column=col).value
        sheet.cell(row=i, column=output_col).value = str(weight_calculator.weight_calc(total)).replace('[', '').replace(
            ']', '')

def update_total(row_min, row_max, col, output_col, sheet, exercise_max):
    for i in range(row_min, row_max):
        multiplier = sheet.cell(row=i, column=col).value
        sheet.cell(row=i, column=output_col).value = multiplier * exercise_max

def main():
    try:
        wb = load_workbook('Workout.xlsx', data_only=True)
    # If the workbook doesn't currently exist, create it
    # TODO: create code to build workbook from scratch if it does not exist
    except FileNotFoundError:
        wb = Workbook()
        wb.save(filename="Workout.xlsx")
    # Create variables to reference to particular sheets in the Excel doc
    maxes = wb['Maxes']
    upper_1 = wb['Upper1']
    lower_1 = wb['Lower1']
    upper_2 = wb['Upper2']
    lower_2 = wb['Lower2']

    user_maxes = {
        "Bench Press": maxes['B2'].value,
        "Overhead Press": maxes['B3'].value,
        "Barbell Row": maxes['B4'].value,
        "Squat": maxes['B5'].value,
        "Deadlift": maxes['B6'].value,
        "Calf Raise": maxes['B7'].value,
    }
    print(user_maxes)



    # Upper1 - Bench Press
    # Update the Excel cells, section by section
    update_total(row_min=3, row_max=10, col=3, output_col=4, sheet=upper_1, exercise_max=user_maxes['Bench Press'])
    iter_weight_calc(row_min=3, row_max=10, col=4, output_col=5, sheet=upper_1)
    # Upper1 - Overhead Press
    update_total(row_min=13, row_max=19, col=3, output_col=4, sheet=upper_1, exercise_max=user_maxes['Overhead Press'])
    iter_weight_calc(row_min=13, row_max=19, col=4, output_col=5, sheet=upper_1)
    # Upper1 - Barbell Row
    update_total(row_min=22, row_max=26, col=3, output_col=4, sheet=upper_1, exercise_max=user_maxes['Barbell Row'])
    iter_weight_calc(row_min=22, row_max=26, col=4, output_col=5, sheet=upper_1)
    # Upper2 - Bench Press
    update_total(row_min=3, row_max=8, col=3, output_col=4, sheet=upper_2, exercise_max=user_maxes['Bench Press'])
    iter_weight_calc(row_min=3, row_max=8, col=4, output_col=5, sheet=upper_2)
    # Upper2 - Close-grip Bench Press
    update_total(row_min=11, row_max=15, col=3, output_col=4, sheet=upper_2, exercise_max=user_maxes['Bench Press'])
    iter_weight_calc(row_min=11, row_max=15, col=4, output_col=5, sheet=upper_2)
    # Upper2 - Barbell Row
    update_total(row_min=18, row_max=22, col=3, output_col=4, sheet=upper_2, exercise_max=user_maxes['Barbell Row'])
    iter_weight_calc(row_min=18, row_max=22, col=4, output_col=5, sheet=upper_2)
    # Lower Days
    # Lower1 - Conventional Squat
    update_total(row_min=3, row_max=10, col=3, output_col=4, sheet=lower_1, exercise_max=user_maxes['Squat'])
    iter_weight_calc(row_min=3, row_max=10, col=4, output_col=5, sheet=lower_1)
    # Lower1 - Sumo Deadlift
    update_total(row_min=13, row_max=17, col=3, output_col=4, sheet=lower_1, exercise_max=user_maxes['Deadlift'])
    iter_weight_calc(row_min=13, row_max=17, col=4, output_col=5, sheet=lower_1)
    # Lower1 - Calf Raise
    update_total(row_min=22, row_max=26, col=3, output_col=4, sheet=lower_1, exercise_max=user_maxes['Calf Raise'])
    iter_weight_calc(row_min=22, row_max=26, col=4, output_col=5, sheet=lower_1)
    # Lower2 - Conventional Deadlift
    update_total(row_min=3, row_max=8, col=3, output_col=4, sheet=lower_2, exercise_max=user_maxes['Deadlift'])
    iter_weight_calc(row_min=3, row_max=8, col=4, output_col=5, sheet=lower_2)
    # Lower2 - Front Squat
    update_total(row_min=12, row_max=18, col=3, output_col=4, sheet=lower_2, exercise_max=user_maxes['Squat'])
    iter_weight_calc(row_min=12, row_max=18, col=4, output_col=5, sheet=lower_2)
    # Lower2 - Calf Raise
    update_total(row_min=21, row_max=25, col=3, output_col=4, sheet=lower_2, exercise_max=user_maxes['Calf Raise'])
    iter_weight_calc(row_min=21, row_max=25, col=4, output_col=5, sheet=lower_2)

    # Save the spreadsheet over the previous spreadsheet
    wb.save('Workout.xlsx')


if __name__ == "__main__":
    main()